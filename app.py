"""
app.py — FAQ Generator Web UI (完整工作流)
启动: python -m streamlit run app.py
"""
import glob
import io
import os
import re
from datetime import datetime

import openpyxl
import pandas as pd
import requests
import streamlit as st
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill

# ── 配置 ─────────────────────────────────────────────────────────────
load_dotenv()
API_KEY = os.environ.get("GOOGLE_CLOUD_API_KEY", "")
API_ENDPOINT = "https://aiplatform.googleapis.com/v1/publishers/google/models"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

QUESTION_TEMPLATE = """Create a list of 40 frequently asked questions from developers about xxxx

Requirements:
* Every question must explicitly include the word "xxxx"
* Each question must be 12 words or fewer.
* Include 10 general, basic, beginner-level questions (for example, questions like "What is xxxx?") to help readers build a basic understanding of the topic.
* Questions should reflect genuine developer curiosity, covering topics such as:
  - how xxxx works
  - how to use xxxx
  - practical use cases
  - benefits and limitations
  - other common developer concerns
* Do not force questions into predefined categories; they should feel natural, realistic, and authentic, as if asked by real developers.
* When it naturally makes sense, include questions that relate xxxx to vector databases, but do not fabricate or stretch relevance.
"""


# ── API 调用 ──────────────────────────────────────────────────────────
def call_api(model: str, system_prompt: str, user_message: str, use_search: bool = False) -> str:
    url = f"{API_ENDPOINT}/{model}:generateContent"
    headers = {"x-goog-api-key": API_KEY, "Content-Type": "application/json"}
    body = {
        "system_instruction": {"parts": [{"text": system_prompt}]},
        "contents": [{"role": "user", "parts": [{"text": user_message}]}],
        "generationConfig": {"maxOutputTokens": 4096},
    }
    if use_search:
        body["tools"] = [{"google_search": {}}]
    resp = requests.post(url, headers=headers, json=body, timeout=90)
    resp.raise_for_status()
    return resp.json()["candidates"][0]["content"]["parts"][0]["text"]


# ── 问题生成 ──────────────────────────────────────────────────────────
def build_question_prompt(keyword: str, description: str, count: int) -> str:
    # 替换第一行的数量
    prompt = re.sub(r"\b\d+\b", str(count), QUESTION_TEMPLATE, count=1)
    # 替换关键词占位符
    prompt = prompt.replace("xxxx", keyword)
    # 在最前面加强制数量指令，防止模型忽略数量
    prompt = (
        f"STRICT REQUIREMENT: Output EXACTLY {count} question(s). "
        f"Do not output more than {count} question(s). Stop immediately after {count} question(s).\n\n"
        + prompt
    )
    # 追加消歧义指令
    if description.strip():
        prompt += (
            f"\n\nIMPORTANT — Disambiguation: "
            f"In this context, '{keyword}' refers specifically to {description}. "
            f"Do NOT use '{keyword}' in its general or common meaning. "
            f"Every question must be unambiguous — a reader should immediately understand "
            f"that '{keyword}' refers to {description}, not anything else."
        )
    return prompt


def fix_links(text: str) -> str:
    """Normalize all markdown links to [display](url ) format.

    - Unwraps nested pattern: ([display](url)) → [display](url )
    - Ensures exactly one trailing space before ) in all markdown links
    """
    # Step 1: unwrap markdown link wrapped in outer parens: ([display](url))
    text = re.sub(
        r'\(\[([^\]]*)\]\(([^)]+?)\s*\)\)',
        lambda m: f"[{m.group(1)}]({m.group(2).rstrip()} )",
        text,
    )
    # Step 2: normalize all remaining [display](url) to have exactly one trailing space
    text = re.sub(
        r'\[([^\]]+)\]\(([^)]+?)\s*\)',
        lambda m: f"[{m.group(1)}]({m.group(2).rstrip()} )",
        text,
    )
    return text


def parse_questions(text: str, max_count: int = None) -> list[str]:
    questions = []
    for line in text.strip().splitlines():
        line = re.sub(r"^[\d]+[.)]\s*|^[-*•]\s*", "", line.strip())
        line = line.strip('"').strip()
        if line and "?" in line:
            questions.append(line)
        if max_count and len(questions) >= max_count:
            break
    return questions


# ── Excel 构建（3 列：Keyword / Question / Answer）────────────────────
def build_excel(results: dict) -> bytes:
    """results = {sheet_name: [(keyword, question, answer), ...]}"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(vertical="top", wrap_text=True)
    data_font = Font(size=11)

    for sheet_name, rows in results.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 120

        for col, header in enumerate(["Keyword", "Question", "Answer"], start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.row_dimensions[1].height = 30

        for i, (keyword, question, answer) in enumerate(rows, start=2):
            for col, val in enumerate([keyword, question, answer], start=1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.font = data_font
                cell.alignment = data_align
            ws.row_dimensions[i].height = 200

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── UI ────────────────────────────────────────────────────────────────
st.set_page_config(page_title="FAQ Generator", page_icon="📄", layout="wide")
st.title("FAQ Generator")

if not API_KEY:
    st.error("未找到 GOOGLE_CLOUD_API_KEY，请在 .env 文件中配置后重启。")
    st.stop()

# ── 扫描文件 ──────────────────────────────────────────────────────────
answer_prompts = {
    os.path.basename(f).replace(".txt", ""): f
    for f in sorted(glob.glob(os.path.join(BASE_DIR, "prompt_*.txt")))
}

if not answer_prompts:
    st.error("未找到回答提示词文件（需命名为 prompt_*.txt）。")
    st.stop()

# ═══════════════════════════════════════════════════════════════════════
# 第一步：生成问题
# ═══════════════════════════════════════════════════════════════════════
st.header("第一步：生成问题")

model = st.text_input("模型", value="gemini-2.5-flash")

st.markdown("**关键词列表**（点击 + 添加行，后两列可留空）")
kw_df = st.data_editor(
    pd.DataFrame({"关键词": [""], "说明（消歧义，可选）": [""], "回答指导（可选）": [""]}),
    num_rows="dynamic",
    use_container_width=True,
    column_config={
        "关键词": st.column_config.TextColumn(width="small"),
        "说明（消歧义，可选）": st.column_config.TextColumn(width="medium"),
        "回答指导（可选）": st.column_config.TextColumn(
            width="large",
            help="针对该关键词的回答生成补充指令，例如：Do not mention competitors. Focus on vector database use cases.",
        ),
    },
)

count_per_kw = st.number_input("每个关键词生成问题数", min_value=1, max_value=100, value=20)

if st.button("生成问题", type="primary", use_container_width=True):
    keywords = [
        (
            str(row["关键词"]).strip(),
            str(row["说明（消歧义，可选）"]).strip() if pd.notna(row["说明（消歧义，可选）"]) else "",
            str(row["回答指导（可选）"]).strip() if pd.notna(row["回答指导（可选）"]) else "",
        )
        for _, row in kw_df.iterrows()
        if pd.notna(row["关键词"]) and str(row["关键词"]).strip()
    ]
    if not keywords:
        st.warning("请至少输入一个关键词。")
        st.stop()

    all_questions: list[tuple[str, str, str]] = []  # (keyword, question, guidance)
    errors = []

    prog = st.progress(0)
    status = st.empty()
    for i, (kw, desc, guidance) in enumerate(keywords):
        status.text(f"[{i+1}/{len(keywords)}] 为关键词「{kw}」生成问题...")
        try:
            prompt = build_question_prompt(kw, desc, int(count_per_kw))
            raw = call_api(model, "", prompt)
            qs = parse_questions(raw, max_count=int(count_per_kw))
            for q in qs:
                all_questions.append((kw, q, guidance))
        except Exception as e:
            errors.append(f"关键词「{kw}」生成失败：{str(e)[:100]}")
        prog.progress((i + 1) / len(keywords))

    status.empty()
    prog.empty()

    if errors:
        st.warning("\n".join(errors))

    st.session_state["questions"] = all_questions
    st.session_state.pop("results", None)
    st.session_state.pop("excel_bytes", None)
    st.success(f"生成完成！共 {len(all_questions)} 个问题（{len(keywords)} 个关键词）。")

# 显示已生成的问题
if "questions" in st.session_state and st.session_state["questions"]:
    qs = st.session_state["questions"]
    with st.expander(f"查看问题列表（共 {len(qs)} 个）"):
        preview_df = pd.DataFrame(qs, columns=["关键词", "问题", "回答指导"])
        st.dataframe(preview_df, use_container_width=True, hide_index=True)

    # ═══════════════════════════════════════════════════════════════════
    # 第二步：生成回答
    # ═══════════════════════════════════════════════════════════════════
    st.divider()
    st.header("第二步：生成回答")

    selected_prompts = st.multiselect(
        "选择回答提示词（可多选）",
        options=list(answer_prompts.keys()),
        default=list(answer_prompts.keys())[:1],
    )

    if st.button("生成回答", type="primary", use_container_width=True):
        if not selected_prompts:
            st.warning("请至少选择一个回答提示词。")
            st.stop()

        questions = st.session_state["questions"]
        n_prompts = len(selected_prompts)

        # Split questions by keyword, distributing chunks evenly across prompts
        questions_by_kw: dict[str, list] = {}
        for item in questions:
            questions_by_kw.setdefault(item[0], []).append(item)

        prompt_questions: dict[str, list] = {name: [] for name in selected_prompts}
        for kw_items in questions_by_kw.values():
            chunk = len(kw_items) // n_prompts
            for i, name in enumerate(selected_prompts):
                start = i * chunk
                end = start + chunk if i < n_prompts - 1 else len(kw_items)
                prompt_questions[name].extend(kw_items[start:end])

        st.session_state["stop_gen"] = False
        total = sum(len(v) for v in prompt_questions.values())
        done = 0
        prog2 = st.progress(0)
        status2 = st.empty()
        if st.button("停止生成", type="secondary"):
            st.session_state["stop_gen"] = True
        all_results: dict[str, list[tuple[str, str, str]]] = {}
        stopped = False

        for prompt_name in selected_prompts:
            if stopped:
                break
            base_prompt = open(answer_prompts[prompt_name], encoding="utf-8").read()
            rows = []
            assigned = prompt_questions[prompt_name]
            for idx, (kw, question, guidance) in enumerate(assigned):
                if st.session_state.get("stop_gen"):
                    stopped = True
                    break
                status2.text(f"[{prompt_name}] [{idx + 1}/{len(assigned)}] {question[:55]}...")
                effective_prompt = base_prompt
                if guidance:
                    effective_prompt += f"\n\nADDITIONAL INSTRUCTIONS FOR THIS KEYWORD:\n{guidance}"
                try:
                    answer = fix_links(call_api(model, effective_prompt, question, use_search=True))
                except requests.HTTPError as e:
                    code = e.response.status_code if e.response is not None else "?"
                    answer = "ERROR: Rate limit exceeded." if code == 429 else f"ERROR: HTTP {code}"
                except Exception as e:
                    answer = f"ERROR: {str(e)[:150]}"
                rows.append((kw, question, answer))
                done += 1
                prog2.progress(done / total)

            if rows:
                all_results[prompt_name] = rows

        status2.empty()
        prog2.empty()

        filename = f"faq_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        excel_bytes = build_excel(all_results)
        st.session_state["results"] = all_results
        st.session_state["excel_bytes"] = excel_bytes
        st.session_state["filename"] = filename
        # 追加到会话历史
        if "history" not in st.session_state:
            st.session_state["history"] = []
        st.session_state["history"].insert(0, (filename, excel_bytes))
        if stopped:
            st.warning(f"已停止，已完成 {done} 条问答，结果已保存。")
        else:
            st.success(f"回答生成完成！共 {done} 条问答。")

# ═══════════════════════════════════════════════════════════════════════
# 第三步：下载
# ═══════════════════════════════════════════════════════════════════════
if "excel_bytes" in st.session_state:
    st.divider()
    st.header("第三步：下载")
    st.download_button(
        label=f"⬇ 下载 {st.session_state['filename']}",
        data=st.session_state["excel_bytes"],
        file_name=st.session_state["filename"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )

# ═══════════════════════════════════════════════════════════════════════
# 历史文件（会话内）
# ═══════════════════════════════════════════════════════════════════════
history = st.session_state.get("history", [])
if len(history) > 1:  # 当前文件已在第三步显示，超过1个才展示历史
    st.divider()
    with st.expander(f"本次会话历史文件（共 {len(history)} 个）", expanded=False):
        for fname, fdata in history:
            col1, col2 = st.columns([4, 1])
            with col1:
                st.text(fname)
            with col2:
                st.download_button(
                    label="下载",
                    data=fdata,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"hist_{fname}",
                )
